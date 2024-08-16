from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import openpyxl
import os
from openpyxl.utils import get_column_letter
class Student:#This class represents a student with their corresponding attributes.
    def __init__(self, name, id, department, section):
        self.name = str(name)  # Ensure name is always a string
        self.id = id
        self.department = department
        self.section = section

    @property
    def surname(self): #Return the student's surname based on the full name.
        return self.name.split()[-1]


class StudentList: # a class for the student list 
    def __init__(self):#Initializes an empty list of students.
        self.students = []

    def add_student(self, student):#Adds selected student object to the list of students.
        self.students.append(student)

    def clear_students(self):#Sets the list of students to an empty list.
        self.students = []

    def get_students_by_section(self, section):#Returns a list of all students in the selected section.
    
        return [student for student in self.students if student.section == section]

class MyApp(Frame):#Main application window which contains various widgets and methods
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.student_list = StudentList()
        self.pack(fill=BOTH, expand=True)
        self.scrollbar=Scrollbar(self,orient=VERTICAL)
        self.scrollbar2=Scrollbar(self,orient=VERTICAL)
        self.label1=Label(self,text="AttendanceKeeper v1.0",font="18")#Title label
        self.label2=Label(self,text="Select student list Excel file: ")
        self.importbutton=Button(self,text="Import List",command=self.clickimport)#create the Import button
        self.label3=Label(self,text="Select a Student:")
        self.label4=Label(self,text="Section:")    
        self.label5=Label(self,text="Attended Students:")
        self.listbox1=Listbox(self,selectmode=MULTIPLE,yscrollcommand=self.scrollbar.set,height=6,width=35)#Create students list boxes with scrollbars attached
        self.var = StringVar(self)
        self.combobox1=ttk.Combobox(self,values=["AP 01","AP 02","AP 03","AP 04","AP 05","AP 06","AP 07","AP 08","AP 09","AP 10","AP 11","AP 12","AP 13","AP 14","AP 15","AP 16","AP 17","AP 18","AP 19","AP 20"],textvariable=self.var)
        self.combobox1.bind("<<ComboboxSelected>>", self.update_student_listbox)#Bind combobox selection event to update student listbox
        self.combobox1.current(0)
        self.filetype = StringVar(self)
        self.combobox2=ttk.Combobox(self,values=[".txt",".csv",".xls"],textvariable=self.filetype,width=7)#create a combobox to select file type
        self.combobox2.current(0)
        self.addbutton=Button(self,text="Add =>",command=self.addstudent)
        self.removebutton=Button(self,text="<= Remove",command=self.removestudent)
        self.exportbutton=Button(self,text="Export as File",command=self.exportfile,width=9)
        self.listbox2=Listbox(self,selectmode=MULTIPLE,yscrollcommand=self.scrollbar2.set,height=6,width=35)#Create the attended list box with scroll bar
        self.label6=Label(self,text="Please select the file type:",width=20)
        self.label7=Label(self,text="Please enter week:")
        self.weekentry=Entry(self,width=17)#Week entry widget
        self.scrollbar.config(command=self.listbox1.yview)
        self.scrollbar.config(command=self.listbox2.yview)
        self.label1.grid(row=0,column=1,sticky=W)
        self.label2.grid(row=1,column=0)
        self.importbutton.grid(row=1,column=1,ipadx=40)
        self.label3.grid(row=2,column=0)
        self.label4.grid(row=2,column=1)
        self.label5.grid(row=2,column=2)
        self.label6.grid(row=3,column=0,sticky=SW)
        self.label7.grid(row=3,column=1,ipady=3,sticky=S)
        self.combobox1.grid(row=3,column=1,sticky=N,ipadx=3.5)
        self.combobox2.grid(row=3,column=0,sticky=SE)
        self.addbutton.grid(row=3,column=1,sticky=N,pady=40,ipadx=50)
        self.removebutton.grid(row=3,column=1,sticky=S,ipadx=40,pady=70)
        self.listbox1.grid(row=3,column=0,sticky=N)
        self.scrollbar.grid(row=3,column=1,sticky=NW,ipady=25)
        self.listbox2.grid(row=3,column=2,sticky=N)
        self.scrollbar2.grid(row=3,column=3,sticky=NE,ipady=25)
        self.weekentry.grid(row=3,column=2,sticky=SW,ipadx=15,ipady=2)
        self.exportbutton.grid(row=3,column=2,sticky=SE)
    def clickimport(self): #Gets filename from user and imports it into program
        self.filepath = filedialog.askopenfilename()
        if self.filepath:
            self.workbook = openpyxl.load_workbook(self.filepath)
            self.AP_studentList = self.workbook.active
            for row in self.AP_studentList.iter_rows(values_only=True):
                student = Student(row[0], row[1], row[2], row[3])
                self.student_list.add_student(student)
            self.update_student_listbox()

    def update_student_listbox(self, event=None):#Updates list box with current students list
        self.listbox1.delete(0, END)
        self.listbox2.delete(0, END) 
        section = self.combobox1.get()  
        students = self.student_list.get_students_by_section(section)
        students.sort(key=lambda x: x.name)  
        for student in students:
            first_name = student.name.split()[0]
            student_data = f" {student.id}, {first_name},{student.surname}"
            self.listbox1.insert(END, student_data)

    def addstudent(self):#Add selected student to attended list box
        self.indexlist=self.listbox1.curselection()
        for index in self.indexlist:
            val = self.listbox1.get(index)
            self.listbox2.insert(END, val)
 
    def removestudent(self):#Remove selected student from attended list box
        self.indexList=self.listbox2.curselection()
        for ind in self.indexList[::-1]:
            self.listbox2.delete(ind)
        
            
    def exportfile(self):#Exports data from the attendance list box to a new file
        section = self.var.get()
        file_type = self.filetype.get()
        file_name = f"{section} {self.weekentry.get()}{file_type}"
        if file_type == ".csv":
            raise BaseException("File type is not supported")

        
        current_directory = os.getcwd()
        if file_type == ".xls": # here when the exported file is excel 
            file_path = os.path.join(current_directory,f"{file_name}" )
            workbook = openpyxl.Workbook()
            worksheet = workbook.active  # Get the active worksheet
            worksheet.title = 'AP_studentlist'   # Assign a title to the worksheet
            worksheet.cell(row=1, column=1, value="ID")
            worksheet.cell(row=1, column=2, value="Name")
            worksheet.column_dimensions[get_column_letter(2)].width = 20
            worksheet.cell(row=1, column=3, value="Department")
            students = self.student_list.get_students_by_section(self.var.get())
            students.sort(key=lambda x: x.name)
            for index, indx in enumerate(students):
                if index in range(self.listbox2.size()):

                    worksheet.cell(row=index+2, column=1,value=indx.name)
                    worksheet.cell(row=index+2, column=2,value=indx.id)
                    worksheet.cell(row=index+2, column=3,value=indx.department) 
                
            workbook.save(file_path) # save the excel file 
        
        elif file_type == ".txt": # here when the exported file is txt 
            file_path=os.path.join(current_directory,f"{file_name}")
            with open(file_path, "w", encoding="utf-8") as file: # use this method to create a new txt 
                students = self.student_list.get_students_by_section(self.var.get())
                students.sort(key=lambda x: x.name)
                file.write("ID\tName\t     Department\n")
                for index, indx in enumerate(students):
                    if index in range(self.listbox2.size()):
                                file.write(f"{indx.name}\t{indx.id}\t{indx.department}\n")


        
def main(): # the main function 
    root = Tk()
    ex = MyApp(root)
    root.geometry("680x280+300+200")
    root.mainloop()
main()